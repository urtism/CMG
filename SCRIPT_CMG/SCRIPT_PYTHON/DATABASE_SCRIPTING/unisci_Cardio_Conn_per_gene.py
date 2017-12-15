import argparse
from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment


class Variante():
	num_alleli_mut_Cardio = 0
	num_alleli_coperti_Cardio = 0
	num_etero_Cardio = 0
	num_HOM_Cardio = 0
	num_paz_mutati_Cardio = 0
	num_paz_tot_Cardio = 0
	paz_list_etero_Cardio = ''
	paz_list_HOM_Cardio = ''

	num_alleli_mut_Conn = 0
	num_alleli_coperti_Conn = 0
	num_etero_Conn = 0
	num_HOM_Conn = 0
	num_paz_mutati_Conn = 0
	num_paz_tot_Conn = 0
	paz_list_etero_Conn = ''
	paz_list_HOM_Conn = ''




def main():
	parser = argparse.ArgumentParser('Unisce le statistiche dei cardio e dei connettivi per il gene indicato')
	
	parser.add_argument('--cardio',help="path del gene nel db per i cardio")
	parser.add_argument('--conn',help="path del gene nel db per i connettivi")
	parser.add_argument('--gene',help="gene inq uestione")
	parser.add_argument('-o','--out',help="file di output senza estensione (senza .tsv o .csv)")

	global opts
	opts = parser.parse_args()

	cardio = open(opts.cardio,'r')
	conn = open(opts.conn,'r')
	varianti = dict()

	for line in cardio:
		line = line.rstrip()
		if line.startswith('CHROM'):
			header = line.split('\t')
		else:
			CHROM,POS,REF,ALT,HGVSc,HGVSp,EXON,INTRON,CONSEQUENCE,FRAZ_ALLELICA,FRAZ_PERC,MAF,ETERO,HOM,NUM_PAZ_MUTATI,PAZIENTI_HET,PAZIENTI_HOM = line.split('\t')
			num_Cardio = int(NUM_PAZ_MUTATI.split('/')[1])
			num_all_Cardio_tot = 2*num_Cardio
			var = Variante()
			var.num_alleli_mut_Cardio = int(FRAZ_ALLELICA.split('/')[0])
			var.num_alleli_coperti_Cardio = int(FRAZ_ALLELICA.split('/')[1])
			var.num_etero_Cardio = int(ETERO)
			var.num_HOM_Cardio = int(HOM)
			var.num_paz_mutati_Cardio = int(NUM_PAZ_MUTATI.split('/')[0])
			num_paz_tot_Cardio = int(NUM_PAZ_MUTATI.split('/')[1])
			if PAZIENTI_HET == '-':
				PAZIENTI_HET = ''
			if PAZIENTI_HOM == '-':
				PAZIENTI_HOM= '' 
			var.paz_list_etero_Cardio = PAZIENTI_HET
			var.paz_list_HOM_Cardio = PAZIENTI_HOM
			varianti['\t'.join([CHROM,POS,REF,ALT,HGVSc,HGVSp,EXON,INTRON,CONSEQUENCE])] = var
	cardio.close()

	for line in conn:
		line = line.rstrip()
		if line.startswith('CHROM'):
			header = line.split('\t')
		else:
			CHROM,POS,REF,ALT,HGVSc,HGVSp,EXON,INTRON,CONSEQUENCE,FRAZ_ALLELICA,FRAZ_PERC,MAF,ETERO,HOM,NUM_PAZ_MUTATI,PAZIENTI_HET,PAZIENTI_HOM = line.split('\t')
			if '\t'.join([CHROM,POS,REF,ALT,HGVSc,HGVSp,EXON,INTRON,CONSEQUENCE]) in varianti.keys():
				var = varianti['\t'.join([CHROM,POS,REF,ALT,HGVSc,HGVSp,EXON,INTRON,CONSEQUENCE])]
			else:
				var = Variante()
			var.num_alleli_mut_Conn = int(FRAZ_ALLELICA.split('/')[0])
			var.num_alleli_coperti_Conn = int(FRAZ_ALLELICA.split('/')[1])
			var.num_etero_Conn = int(ETERO)
			var.num_HOM_Conn = int(HOM)
			var.num_paz_mutati_Conn = int(NUM_PAZ_MUTATI.split('/')[0])
			num_paz_tot_Conn = int(NUM_PAZ_MUTATI.split('/')[1])
			if PAZIENTI_HET == '-':
				PAZIENTI_HET = ''
			if PAZIENTI_HOM == '-':
				PAZIENTI_HOM= '' 
			var.paz_list_etero_Conn = PAZIENTI_HET
			var.paz_list_HOM_Conn = PAZIENTI_HOM
			varianti['\t'.join([CHROM,POS,REF,ALT,HGVSc,HGVSp,EXON,INTRON,CONSEQUENCE])] = var
	conn.close()

	out = open(opts.out + '.tsv','w')
	out.write('\t'.join(['CHROM','POS','REF','ALT','HGVSc','HGVSp','EXON','INTRON','CONSEQUENCE','SYMBOL','FRAZ_ALLELICA','FRAZ_PERC','MAF','ETERO','HOM','NUM_PAZ_MUTATI','PAZIENTI_HET','PAZIENTI_HOM'])+'\n')

	wb = Workbook()
	dest_filename = opts.out + '.xlsx'
	ws1 = wb.active
	ws1.title = 'statistiche ' + opts.gene
	ws1.append(['CHROM','POS','REF','ALT','HGVSc','HGVSp','EXON','INTRON','CONSEQUENCE','SYMBOL','FRAZ_ALLELICA','FRAZ_PERC','MAF','ETERO','HOM','NUM_PAZ_MUTATI','PAZIENTI_HET','PAZIENTI_HOM'])

	for id in varianti.keys():
		var = varianti[id]

		num_alleli_mut = var.num_alleli_mut_Cardio + var.num_alleli_mut_Conn

		if var.num_alleli_coperti_Cardio == 0:
			num_alleli_coperti = (2*num_paz_tot_Cardio) + var.num_alleli_coperti_Conn

		elif var.num_alleli_coperti_Conn == 0:
			num_alleli_coperti = var.num_alleli_coperti_Cardio + (2*num_paz_tot_Conn)

		else:
			num_alleli_coperti = var.num_alleli_coperti_Cardio + var.num_alleli_coperti_Conn

		num_etero = var.num_etero_Cardio + var.num_etero_Conn
		num_HOM = var.num_HOM_Cardio + var.num_HOM_Conn
		num_paz_mutati = var.num_paz_mutati_Cardio + var.num_paz_mutati_Conn
		num_paz_tot = num_paz_tot_Cardio + num_paz_tot_Conn
		paz_list_etero = ';'.join([var.paz_list_etero_Cardio,var.paz_list_etero_Conn])
		if paz_list_etero == ';' or paz_list_etero == '-;':
			paz_list_etero = '-' 
		paz_list_HOM = ';'.join([var.paz_list_HOM_Cardio,var.paz_list_HOM_Conn])
		if paz_list_HOM == ';' or paz_list_HOM == '-;':
			paz_list_HOM = '-' 


		FRAZ_ALLELICA = str(num_alleli_mut) + '/' + str(num_alleli_coperti) 
		FRAZ_PERC = str(round((float(num_alleli_mut)/float(num_alleli_coperti))*100.0,2))+'%'
		MAF = str(round(float(num_alleli_mut)/float(num_alleli_coperti),4))
		ETERO = str(num_etero)
		HOM = str(num_HOM)
		NUM_PAZ_MUTATI = str(num_paz_mutati) + '/' +str(num_paz_tot)
		PAZIENTI_HET = paz_list_etero
		PAZIENTI_HOM = paz_list_HOM
		GENE = opts.gene

		out.write('\t'.join([id,GENE,FRAZ_ALLELICA,FRAZ_PERC,MAF,ETERO,HOM,NUM_PAZ_MUTATI,PAZIENTI_HET,PAZIENTI_HOM])+'\n')
		to_print = id.split('\t') + [GENE,FRAZ_ALLELICA,FRAZ_PERC,MAF,ETERO,HOM,NUM_PAZ_MUTATI,PAZIENTI_HET,PAZIENTI_HOM]
		ws1.append(to_print)

	for col in ws1.columns:
		for cell in col:
			cell.alignment = Alignment(horizontal="center", vertical="center")
	wb.save(filename = dest_filename)
main()