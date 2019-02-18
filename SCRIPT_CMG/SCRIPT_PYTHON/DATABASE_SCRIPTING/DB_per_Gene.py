import os
import glob2
import argparse
import shutil
from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment

def extract_info(varianti,hash_var,pazienti):
	index_of_gene = 0
	index_of_gen = 0
	buff = []
	for line in varianti:
		line = line.rstrip()
		if line.startswith('CHROM'):
			#print line
			index_of_gene = line.split('\t').index('SYMBOL')
			index_of_gen = line.split('\t').index('GT')
			index_of_ac = line.split('\t').index('AC')
			index_of_an = line.split('\t').index('AN')
			index_of_exon = line.split('\t').index('EXON')
			index_of_intron = line.split('\t').index('INTRON')

			#print line
			
			try:
				index_of_HGVSc = line.split('\t').index('HGVSc')
			except:
				print "non trova hgsvc",line
			
			try:
				index_of_HGVSp = line.split('\t').index('HGVSp')
			except:
				print "non trova hgsvp",line
			try:
				index_of_conseq = line.split('\t').index('Consequence')
			except:
				print "non trova CONSEQUENCE",line

		else:
			variante = line.split('\t')
			chrom = variante[0]
			pos = variante[1]
			id = variante[2]
			ref = variante[3]
			alt = variante[4]
			HGVSc = variante[index_of_HGVSc] 
			HGVSp = variante[index_of_HGVSp]
			CONSEQUENCE = variante[index_of_conseq]
			AC = variante[index_of_ac]
			AN = variante[index_of_an]
			INTRON = variante[index_of_intron]
			EXON = variante[index_of_exon]
			info = variante[4:]
			var_id = '\t'.join([chrom,pos,ref,alt])


			if var_id in hash_var.keys():
				try:

					if variante[index_of_gen].split('/')[0] == variante[index_of_gen].split('/')[1]:
						#print "gia presente",hash_var[var_id],AC,AN
						hash_var[var_id][1] = hash_var[var_id][1] + 1
						try:
							hash_var[var_id][2][1] = hash_var[var_id][2][1] + [pazienti[id][0]]
						except:
							hash_var[var_id][2][1] = hash_var[var_id][2][1] + [id]
					else: 
						
						hash_var[var_id][0] = hash_var[var_id][0] + 1

						try:
							hash_var[var_id][2][0] = hash_var[var_id][2][0] + [pazienti[id][0]]
						except:
							hash_var[var_id][2][0] = hash_var[var_id][2][0] + [id]
				except:
					print variante

				hash_var[var_id][3] = HGVSc
				hash_var[var_id][4] = HGVSp
				hash_var[var_id][5] = CONSEQUENCE
				hash_var[var_id][6] = AC
				hash_var[var_id][7] = AN
				hash_var[var_id][8] = EXON
				hash_var[var_id][9] = INTRON


			else:
				try:
					if variante[index_of_gen].split('/')[0] == variante[index_of_gen].split('/')[1]:
						try:
							hash_var[var_id] = [0,1,[[],[pazienti[id][0]]],HGVSc,HGVSp,CONSEQUENCE,AC,AN,EXON,INTRON]
						except:
							hash_var[var_id] = [0,1,[[],[id]],HGVSc,HGVSp,CONSEQUENCE,AC,AN,EXON,INTRON]
								
					else:
						try:
							hash_var[var_id] = [1,0,[[pazienti[id][0]],[]],HGVSc,HGVSp,CONSEQUENCE,AC,AN,EXON,INTRON]
						except:
							hash_var[var_id] = [0,1,[[id],[]],HGVSc,HGVSp,CONSEQUENCE,AC,AN,EXON,INTRON]
				except:
					print variante
				#print "nuova variante",hash_var[var_id],AC,AN


def extract_var_from_gene(varianti):
	var_list = open(opts.out + '/lista_varianti_' + opts.gene+ '.list','a+')
	var_array = var_list.readlines()
	index_of_gene = 0
	
	for line in varianti:
		#print line
		line = line.rstrip()
		line_split= line.split('\t')
		if line.startswith('CHROM'):
			
			var_id = line_split[0:5]
			index_of_gen = line_split.index('GT')
			index_of_ac = line_split.index('AC')
			index_of_an = line_split.index('AN')
			index_of_HGVSc = line_split.index('HGVSc')
			index_of_HGVSp = line_split.index('HGVSp')
			index_of_conseq = line_split.index('Consequence')
			index_of_gene = line_split.index('SYMBOL')
			index_of_exon = line_split.index('EXON')
			index_of_intron = line_split.index('INTRON')


			header = '\t'.join(var_id + ['HGVSc','HGVSp','Consequence','SYMBOL','GT','AC','AN','EXON','INTRON']) + '\n'
			if header not in var_array: 
				var_list.write(header)
			
		else:
			if line_split[index_of_gene].rstrip() == opts.gene.rstrip():
				#print "si"
				var_id = line_split[0:5]
				var_list.write('\t'.join(var_id + [line_split[index_of_HGVSc],line_split[index_of_HGVSp],line_split[index_of_conseq],line_split[index_of_gene],line_split[index_of_gen],line_split[index_of_ac],line_split[index_of_an],line_split[index_of_exon],line_split[index_of_intron]]) + '\n')
			
def extract_paz_name(pazienti):
	id_info={}
	for paz in pazienti:
		paz = paz.rstrip()
		if paz.startswith('NOME_PAZ') or paz == '':
			continue
		else:
			#print paz
			try:
				nome = paz.split('\t')[0]
				num_paz = paz.split('\t')[1]
				num_run = paz.split('\t')[2]
				data_run = paz.split('\t')[3]
				id = paz.split('\t')[4]
				id_info[id]=[nome,num_paz,num_run,data_run]
			except:
				print "il paziente",paz.split('\t')[0],"ha un formato non valido"
	return id_info		

def main():
	parser = argparse.ArgumentParser('Adds number of exon for each interval')
	
	parser.add_argument('-p','--path',help="path dei tsv")
	parser.add_argument('-g','--gene',help="gene to search")
	parser.add_argument('-o','--out',help="path di output")
	parser.add_argument('-l','--paz_list',help="lista dei nomi dei pazienti e degli id")
	parser.add_argument('-t','--tipo',help="Cancer o Cardio",default='Cardio')

	global opts
	opts = parser.parse_args()
	global pazienti
	opts.out=opts.out.rstrip()
	opts.gene=opts.gene.rstrip()

	pazienti = extract_paz_name(open(opts.paz_list,'r'))
	
	try:
		os.mkdir(opts.out)
	except:
		#print opts.out,"directory esistente"
		shutil.rmtree(opts.out)
		os.mkdir(opts.out)

	hash_var = {}
	num_paz = 0

	for filename in glob2.glob(os.path.join(opts.path,'*.tsv')):
		num_paz = num_paz + 1

		extract_var_from_gene(open(filename,'r'))
	
	#print num_paz

	for filename in glob2.glob(os.path.join(opts.out,'*.list')):
		if 'GATK_Other' in filename:
			continue
		else:
			extract_info(open(filename,'r'),hash_var,pazienti)
		#os.remove(filename)

	statistiche = open(opts.out + '/lista_varianti_' + opts.gene + '_STATS.tsv','w')
	wb = Workbook()
	dest_filename = opts.out + '/lista_varianti_' + opts.gene + '_STATS.xlsx'
	ws1 = wb.active
	ws1.title = "range names"


	statistiche.write('CHROM'+'\t' + 'POS'+'\t' + 'REF'+'\t' + 'ALT' +'\t' +'HGVSc'+'\t' + 'HGVSp'+'\t'+'EXON'+'\t'+'INTRON'+'\t'+'CONSEQUENCE'+'\t'+ 'FRAZ_ALLELICA'+'\t' + 'FRAZ_PERC'+'\t' + 'MAF'+'\t'+'ETERO'+'\t'+ 'HOM'+'\t'+'NUM_PAZ_MUTATI'+'\t'+'PAZIENTI_HET'+'\t'+'PAZIENTI_HOM'+'\n')
	header=['CHROM','POS','REF','ALT','HGVSc','HGVSp','EXON','INTRON','CONSEQUENCE','FRAZ_ALLELICA','FRAZ_PERC','MAF','ETERO','HOM','NUM_PAZ_MUTATI','PAZIENTI_HET','PAZIENTI_HOM']
	ws1.append(header)

	for var in hash_var.keys():
		if hash_var[var][2][0] ==[]:
			hash_var[var][2][0] =['-']
		if hash_var[var][2][1] ==[]:
			hash_var[var][2][1] =['-']
		maf = float("{0:.4f}".format(float(hash_var[var][6])/float(hash_var[var][7])))
		fraz = str(hash_var[var][6]) + '/' + str(hash_var[var][7])
		perc = str(maf*100) + '%'
		paz_mut = str(int(hash_var[var][0]) + int(hash_var[var][1])) + '/' + str((num_paz))
		info_var = str(hash_var[var][3]) + '\t' + str(hash_var[var][4])+  '\t' + str(hash_var[var][8])+ '\t' + str(hash_var[var][9]) +'\t' + str(hash_var[var][5])
		stats = fraz + '\t' + perc + '\t' +  str(maf) + '\t' + str(hash_var[var][0]) + '\t' + str(hash_var[var][1]) + '\t' + str(paz_mut) + '\t' + ';'.join(hash_var[var][2][0])+ '\t' + ';'.join(hash_var[var][2][1])
		statistiche.write(var + '\t' +info_var + '\t' + stats + '\n')

		var_split1 = var.split('\t')
		info_var1 = [str(hash_var[var][3]),str(hash_var[var][4]),str(hash_var[var][8]),str(hash_var[var][9]),str(hash_var[var][5])]
		stats1 = [fraz,perc, str(maf),str(hash_var[var][0]),str(hash_var[var][1]),str(paz_mut),';'.join(hash_var[var][2][0]),';'.join(hash_var[var][2][1])]
		to_print=var_split1+info_var1+stats1
		ws1.append(to_print)
	
	for col in ws1.columns:
		for cell in col:
			cell.alignment = Alignment(horizontal="center", vertical="center")
	wb.save(filename = dest_filename)

main()
